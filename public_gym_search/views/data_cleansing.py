import traceback
import re
from django.db import transaction
from openpyxl import load_workbook

from const import Const
from public_gym_search.models.models import PublicGymnasium, create_primary_key


class GymExcelModel(object):
    """ スクレイピング結果→Excel """

    def __init__(self):
        self.index = Const.INT_UNSET  # for文のindex用

        self.facility_name = Const.STRING_EMPTY
        self.prefecture = Const.STRING_EMPTY
        self.municipality = Const.STRING_EMPTY
        self.address = Const.STRING_EMPTY
        self.telephone = None
        self.url = None
        self.training_room_flag = None


class DataCleansing(object):
    """ Data Cleansing """

    def __init__(self, data_list=None):
        self.data_list = data_list  # スクレイピング結果オブジェクト
        self.wb = load_workbook("public_gym_list.xlsx")  # Excel Book
        self.ws = self.wb["Sheet1"]  # Excel Sheet

    def get_value(self, row, column):
        """ 対象セルの値を取得する """
        if column < 0 or row < 0:
            return ""

        value = str(self.ws.cell(row=row, column=column).value)

        if Const.is_null_or_empty(value):
            return ""

        return value

    def set_value(self, row, column, value):
        """ 対象セルに値をセットする """
        self.ws.cell(row=row, column=column).value = value

    def address_to_address(self, value):
        """ 取得した生addressを加工してきれいにする。
        変更前：「所在地： 〒102-0091 東京都千代田区北の丸公園２－３」
        変更後：「東京都千代田区北の丸公園２－３」
        ※本処理は完璧でない。ファイルを確認し、修正されていないデータを手動で修正する
        """
        pattern = r"所在地： 〒\d{3}-\d{4}\s"
        result = re.sub(pattern, "", value)

        return result

    def get_prefecture_and_municipality_from_address(self, value):
        """
        所在地から都道府県と市区町村を抽出する
        ※本処理は完璧でない。ファイルを確認し、修正されていないデータを手動で修正する
        """
        pattern = """(...??[都道府県])((?:旭川|伊達|石狩|盛岡|奥州|田村|南相馬|那須塩原|東村山|武蔵村山|羽村|十日町|上越|
        富山|野々市|大町|蒲郡|四日市|姫路|大和郡山|廿日市|下松|岩国|田川|大村|宮古|富良野|別府|佐伯|黒部|小諸|塩尻|玉野|
        周南)市|(?:余市|高市|[^市]{2,3}?)郡(?:玉村|大町|.{1,5}?)[町村]|(?:.{1,4}市)?[^町]{1,4}?区|.{1,7}?[市町村])(.+)"""

        result = re.match(pattern, value)

        return result

    def db_insert(self):
        """ Excel to DB """
        self.wb = load_workbook("public_gym_list.xlsx")
        self.ws = self.wb["Sheet1"]

        # Excel-Dataを取得し、Modelリストを作成する
        model_list = []
        for row in range(2, self.ws.max_row + 1):
            model = GymExcelModel()
            model.facility_name = self.get_value(row, 1)
            model.prefecture = self.get_value(row, 2)
            model.municipality = self.get_value(row, 3)
            model.address = self.get_value(row, 4)
            model_list.append(model)
        print(model_list)
        print(len(model_list))

        print("モデルリストの作成が終了した。")

        # Model → DB INSERT
        count = 0
        models = []
        try:
            with transaction.atomic():
                for model in model_list:
                    public_gymnasium = PublicGymnasium(
                        id=create_primary_key(),
                        facility_name=model.facility_name,
                        prefecture=model.prefecture,
                        municipality=model.municipality,
                        address=model.address,
                        telephone=model.telephone,
                        url=model.url,
                        training_room_flag=model.training_room_flag
                        )
                    models.append(public_gymnasium)
                    count += 1

                PublicGymnasium.objects.bulk_create(models)

        except Exception as e:
            print(e)
            print(traceback.format_exc())
        finally:
            self.wb.close()

        print(f"{ count }件のデータをINSERTしました。")

    def data_cleansing_process(self):
        """ データクレンジング処理 """
        model_list = []

        for i, element in enumerate(self.data_list):
            for j, sublist in enumerate(element):
                model = GymExcelModel()
                for k, item in enumerate(sublist):
                    if k == 0:
                        model.facility_name = item
                    if k == 1:
                        model.address = self.address_to_address(item)
                else:
                    model_list.append(model)

        print(len(model_list))

        # Excel 書込み
        for i, model in enumerate(model_list):
            i += 2
            self.set_value(i, 1, model.facility_name)
            self.set_value(i, 4, model.address)

        print("処理が完了しました")
        self.wb.save("public_gym_list.xlsx")
        self.wb.close()

    def data_cleansing_address(self):
        # ここからはExcelファイルを再開封し直接編集する
        self.wb = load_workbook("public_gym_list.xlsx")
        self.ws = self.wb["Sheet1"]

        model_list = []
        for row in range(2, self.ws.max_row + 1):
            model = GymExcelModel()
            model.facility_name = self.get_value(row, 1)
            model_list.append(model)

        print("箱完成")

        failure_list = []
        for i, model in enumerate(model_list):
            model.index = i
            i += 2
            model.address = self.get_value(i, 4)
            prefecture_and_municipality = \
                self.get_prefecture_and_municipality_from_address(model.address)
            if prefecture_and_municipality:
                model.prefecture = prefecture_and_municipality.group(1)
                model.municipality = prefecture_and_municipality.group(2)
            else:
                failure_list.append(model)
                model.index = Const.INT_UNSET

        print(len(failure_list))

        print("Excelから値を取得し、モデルに格納した。")

        for model in model_list:
            if model.index == -1:
                continue
            row = model.index + 2
            self.set_value(row, 2, model.prefecture)
            self.set_value(row, 3, model.municipality)

        self.wb.save("public_gym_list.xlsx")
        self.wb.close()
        print("処理が終了しました。")
