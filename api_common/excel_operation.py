"""Excel Operation module"""
from openpyxl import load_workbook

from const import Const


class ExcelOperation(object):
    """Base Excel Operation"""

    def __init__(self):
        """コンストラクタ"""
        pass

    def __del__(self):
        """デストラクタ"""
        print("ExcelOperationオブジェクトを破棄します。")

    # ===== Common Parts =========================================
    def read_excel(self, workbook, worksheet):
        """read excel book and sheet

        Args:
            workbook (str): 対象のExcelBookのPath
            worksheet (str): 対象のExcelSheet名
        """
        workbook = load_workbook(workbook)
        worksheet = workbook[worksheet]

        return workbook, worksheet

    def get_cell_value(self, worksheet, row, column) -> str:
        """get cell value

        Args:
            worksheet (): 対象のExcelBookのシートオブジェクト
            row (int): 取得する値のセルの行番号
            column (int): 取得する値のセルの列番号
        """
        if row < 0 or column < 0:
            return ""

        value = str(worksheet.cell(row=row, column=column).value)

        if Const.is_null_or_empty(value):
            return ""

        return value

    def set_cell_value(self, worksheet, row, column, value):
        """set cell value

        Args:
            worksheet (): 対象のExcelBookのシートオブジェクト
            row (int): 入力したいセルの行番号
            column (int): 入力したいセルの列番号
            value (): 入力する値
        """
        worksheet.cell(row=row, column=column).value = value

    def save_excel(self, workbook, file_path=None):
        """Save ExcelBook and Close"""
        workbook.save(file_path)

    def save_and_close_excel(self, workbook, file=None, save_flag=False):
        """Save ExcelBook and Close"""
        if save_flag:
            workbook.save(file)

        workbook.close()

    def excel_to_pdf(self):
        """Excel to PDF"""
        pass
